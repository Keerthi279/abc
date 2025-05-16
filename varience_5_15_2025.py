import polars as pl
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, NamedStyle, Color
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, Tuple, Union
from functools import cached_property
import glob
import re
from itertools import product
from datetime import datetime
from functools import reduce
from collections import defaultdict
from joblib import Parallel, delayed
from openpyxl.utils.dataframe import dataframe_to_rows



 
 
class MetaSingleton(type):
    """Classic MetaSingleton creates an instance only if there is no instance created so far;
    otherwise, it will return the instance that is already created."""
 
    _instances = {}
    def __call__(cls, *args, **kwargs):
        if cls not in cls._instances:
            cls._instances[cls] = super().__call__(*args, **kwargs)
        return cls._instances[cls]
 
@dataclass(frozen=True)
class RubyProductionData(metaclass=MetaSingleton):
    input_data_location: Path
    scenario_data_location: Path
    columns_for_sorting: Tuple[str] = field(
        default_factory=lambda: ["SCN_ID", "CURVE_DETL_ID", "CURVE_MSTR_ID", "CURVE_TENOR", "AS_OF_DATE"]
    )
 
    @cached_property
    def scen_data_polars(self):
        df = pl.read_excel(self.scenario_data_location, engine="openpyxl").select(
            [pl.col("SCN_ID").alias("SCN_ID_right"), pl.col("SCN_NAME")] ## TODO: need to validate that these column headers are present in the Dataframe
        )
       
        return df
 
    @cached_property
    def input_data_polars(self):
        ## TODO: need to validate that these column headers are present in the Dataframe
        df = (
            pl.read_parquet(
                self.input_data_location,
            ).sort(
                self.columns_for_sorting, descending=False
            ).with_columns(
                pl.col("CURVE_TENOR").replace_strict(
                    {"ZERO": 0, "O/N": 1, "7D": 7, "14D": 14, **{f"{i}M": i*30 for i in range(1, 600)}}
                ).cast(pl.Int32).alias("curve_tenor_days"),
            ).join(
                self.scen_data_polars,
                left_on = "SCN_ID", right_on = "SCN_ID_right",
            )
        )
       
        ## TODO: need to validate that these column headers are present in the Dataframe
        new_cols_order = ["SCN_ID", "SCN_NAME", "AS_OF_DATE", "CURVE_CODE", "CCY_CODE", "curve_tenor_days", "CURVE_TENOR"] + \
            [f"MON{i}" for i in range(-24, 61)] + \
            ["CURVE_DETL_ID", "CURVE_MSTR_ID", "DATASET_ID", "CURVE_DESC", "UNITS", "ORIGL_FLAG", "CURVE_FLOOR", "CURVE_CNTRY", "FWD_SPRD_BOOL",
            "CREATED_BY", "CREATE_TS", "SOURCE", "SOURCE_DETAILS"]
        df = df.select(new_cols_order)
       
        return df
 
def truncate_sheet_name_length( original_str: str, field_to_truncate: str, max_length: int = 31) -> str:
    if len(original_str) <= max_length:
        return original_str # if less than max_length, just return original str
    _remainder_str = original_str.replace(field_to_truncate, '')
    _residual_len = max_length - len(_remainder_str)
 
    return re.sub(f"({re.escape(field_to_truncate)})", field_to_truncate[:_residual_len], original_str)

def compute_diff_for_combination(scn, curve, ccy, df, curr_date, prior_date, forward_cols, abs_tol):
    filter_expr = (
        (pl.col("SCN_ID") == scn) &
        (pl.col("CURVE_CODE") == curve) &
        (pl.col("CCY_CODE") == ccy)
    )
    curr_df = df.filter(filter_expr & (pl.col("AS_OF_DATE") == curr_date)).with_columns(
        *[pl.col(c) * 100 for c in forward_cols]
    )
    prior_df = df.filter(filter_expr & (pl.col("AS_OF_DATE") == prior_date)).with_columns(
        *[pl.col(c) * 100 for c in forward_cols]
    )

    if curr_df.height == 0 or prior_df.height == 0 or curr_df.shape != prior_df.shape:
        return None

    curr_pd, prior_pd = curr_df.to_pandas(), prior_df.to_pandas()
    offset = 1
    shifted_curr_cols = forward_cols[:-offset]
    shifted_prior_cols = forward_cols[offset:]
    index_cols = ["SCN_ID", "CURVE_CODE", "CCY_CODE", "CURVE_TENOR", "SCN_NAME"]
    diff_values = np.abs(curr_pd[shifted_curr_cols].values - prior_pd[shifted_prior_cols].values)

    abs_diff = pd.DataFrame(
        diff_values,
        columns=shifted_curr_cols
    )
    abs_diff[index_cols] = curr_pd[index_cols]
    abs_diff = abs_diff.loc[(abs_diff[shifted_curr_cols] > abs_tol).any(axis=1)]

    if abs_diff.empty:
        return None

    return {
        "scn": scn,
        "curve": curve,
        "ccy": ccy,
        "curr": curr_pd,
        "prior": prior_pd,
        "diff": abs_diff.reset_index()
    }

def write_variance_report_to_excel(results, output_path):
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "VARIANCE_SUMMARY"

    summary_rows = []
    for res in results:
        if res is None:
            continue
        scen, curve, ccy = res["scn"], res["curve"], res["ccy"]
        count_exceed = res["diff"].shape[0]
        summary_rows.append([scen, curve, ccy, count_exceed])

    ws_summary.append(["SCN_ID", "CURVE_CODE", "CCY_CODE", "Records_Above_Tolerance"])
    for row in summary_rows:
        ws_summary.append(row)

    for res in results:
        if res is None:
            continue
        scen, curve, ccy = res["scn"], res["curve"], res["ccy"]
        sheet_base = f"{scen}_{curve}_{ccy}"[:31]

        ws_stacked = wb.create_sheet(title=f"{sheet_base}_Stacked"[:31])
        stacked = pd.concat([
            res["curr"].assign(Source="Current Month"),
            res["prior"].assign(Source="Prior Month")
        ])
        for r in dataframe_to_rows(stacked, index=False, header=True):
            ws_stacked.append(r)

        ws_diff = wb.create_sheet(title=f"{sheet_base}_ABS_DIFF"[:31])
        for r in dataframe_to_rows(res["diff"], index=False, header=True):
            ws_diff.append(r)

    wb.save(output_path)

def process_combination(scen, curve, ccy, _df, CURR_DATE, PRIOR_DATE, _forward_months_cols, abs_tolerance_in_pct):
    _curve_identifiers = ["SCN_ID", "CURVE_CODE", "CCY_CODE"]
    _metadata_Cols = ["CURVE_MSTR_ID", "SCN_NAME", "ORIGL_FLAG", "CURVE_FLOOR", "CREATED_BY", "CREATE_TS", "SOURCE", "SOURCE_DETAILS"]

    curr_df = _df.filter(
        (pl.col("SCN_ID") == scen) & (pl.col("CURVE_CODE") == curve) & (pl.col("CCY_CODE") == ccy) & (pl.col("AS_OF_DATE") == CURR_DATE)
    ).with_columns(
        *[pl.col(i) * 100 for i in _forward_months_cols]
    ).select(["DATASET_ID", "AS_OF_DATE"] + _curve_identifiers + ["CURVE_TENOR"] + _forward_months_cols + _metadata_Cols)

    prior_df = _df.filter(
        (pl.col("SCN_ID") == scen) & (pl.col("CURVE_CODE") == curve) & (pl.col("CCY_CODE") == ccy) & (pl.col("AS_OF_DATE") == PRIOR_DATE)
    ).with_columns(
        *[pl.col(i) * 100 for i in _forward_months_cols]
    ).select(["DATASET_ID", "AS_OF_DATE"] + _curve_identifiers + ["CURVE_TENOR"] + _forward_months_cols + _metadata_Cols)

    if curr_df.shape != prior_df.shape:
        return None

    curr_pd = curr_df.to_pandas()
    prior_pd = prior_df.to_pandas()

    offset = 1
    shifted_prior_cols = _forward_months_cols[offset:]
    shifted_curr_cols = _forward_months_cols[:-offset]
    _abs_diff_idx = ["DATASET_ID"] + _curve_identifiers + ["CURVE_TENOR", "SCN_NAME"]

    _abs_diff_df = pd.DataFrame(
        np.abs(curr_pd[shifted_curr_cols].values - prior_pd[shifted_prior_cols].values),
        index=curr_pd[_abs_diff_idx].values,
        columns=shifted_curr_cols,
    )
    _abs_diff_df.index = pd.MultiIndex.from_tuples(_abs_diff_df.index, names=_abs_diff_idx)

    if (_abs_diff_df.values > abs_tolerance_in_pct).any():
        return {
            "scn": scen,
            "curve": curve,
            "ccy": ccy,
            "curr": curr_pd,
            "prior": prior_pd,
            "diff": _abs_diff_df
        }
    return None

def optimized_reader_export(_df:RubyProductionData, scenario_ids, curve_codes, ccy_codes, output_filename, abs_tolerance_in_pct=0.7):
    _forward_months_cols = [f"MON{i}" for i in range(0, 61)]
    _dates_used = pd.to_datetime(_df.select("AS_OF_DATE").unique().to_pandas().values.flatten())
    CURR_DATE, PRIOR_DATE = max(_dates_used), min(_dates_used)

    # Parallel processing for all combinations
    tasks = list(product(scenario_ids, curve_codes, ccy_codes))
    results = Parallel(n_jobs=8)(
        delayed(process_combination)(
            scen, curve, ccy, _df, CURR_DATE, PRIOR_DATE, _forward_months_cols, abs_tolerance_in_pct
        ) for scen, curve, ccy in tasks
    )

    results = [res for res in results if res]

    # Create initial workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "VARIANCE_SUMMARY"
    wb.save(output_filename)
    write_variance_report_to_excel(results, "output/custom.xlsx")


    for res in results:
        scen, curve, ccy = res["scn"], res["curve"], res["ccy"]
        curr_pd, prior_pd = res["curr"], res["prior"]
        abs_diff_df = res["diff"]

        with pd.ExcelWriter(output_filename, engine="openpyxl", mode='a', if_sheet_exists='new') as writer:
            sheet_base = truncate_sheet_name_length(f"{scen}_{curve}_{ccy}", curve, max_length=31)
            combined_df = pd.concat(
                [curr_pd.assign(Source="Current Month"), prior_pd.assign(Source="Prior Month")],
                ignore_index=True
            )
            combined_df.to_excel(writer, sheet_name=f"{sheet_base}_Stacked", index=False)
            abs_diff_df.reset_index().to_excel(writer, sheet_name=f"{sheet_base}_ABS_DIFF", index=False)

    return output_filename


def reader_and_export(
    ruby_data: RubyProductionData,
    scenario_id: Union[int, Tuple[int]],
    curve_code: Union[str, Tuple[str]],
    ccy_code: Union[str, Tuple[str]],
    output_filename: Path,
    abs_tolerance_in_pct: float = 0.7,
    forward_months_cols: Tuple[str] = [f"M{i}" for i in range(0, 61)],
    find_base_vs_shock: bool =False
) -> None:
    
    _forward_months_cols = ["MON{used_int}".format(used_int=re.findall(r'([\-|\+]*\d+)', i)[0]) for i in forward_months_cols]
    _scenario_ids = [scenario_id] if isinstance(scenario_id, int) else list(scenario_id)
    _curve_codes = [curve_code] if isinstance(curve_code, str) else list(curve_code)
    _ccy_codes = [ccy_code] if isinstance(ccy_code, str) else list(ccy_code)
    _df = ruby_data.input_data_polars
   
    _dates_used = sorted(_df.select("AS_OF_DATE").unique().to_series().to_list())
    assert len(_dates_used) == 2, "Make sure there are only current & prior months!"
    CURR_DATE, PRIOR_DATE = max(_dates_used), min(_dates_used)
    combos = list(product(_scenario_ids, _curve_codes, _ccy_codes))
    results = Parallel(n_jobs=4)(
        delayed(compute_diff_for_combination)(
            scn, curve, ccy, _df, CURR_DATE, PRIOR_DATE, _forward_months_cols, abs_tolerance_in_pct
        ) for scn, curve, ccy in combos
    )
    results = [r for r in results if r]
    write_variance_report_to_excel(results, "output/custom.xlsx")


    _curve_identifiers = ["SCN_ID", "CURVE_CODE", "CCY_CODE"]
    _metadata_Cols = ["CURVE_MSTR_ID", "SCN_NAME", "ORIGL_FLAG", "CURVE_FLOOR", "CREATED_BY", "CREATE_TS", "SOURCE", "SOURCE_DETAILS"]
    _combined_checks_above_tol_list = []
    if output_filename.exists(): ## remove the output file if it already exists # TODO: Need to check if it is being opened
        os.remove(output_filename)
       
    wb = Workbook()
    ws = wb.active
    ws.title = "VARIANCE_SUMMARY"
 
 
    # Save the workbook
    wb.save(output_filename)
 
    print("Excel file with multiple empty sheets created successfully!")
    for scen, curve, ccy in product(*[_scenario_ids, _curve_codes, _ccy_codes]):
        row = {"SCN_ID": scen, "CURVE_CODE": curve, "CCY_CODE": ccy} # TODO: refactor this unnecessary assignment
        ## TODO: There's code smell everywhere, need to refactor when there's less time-pressure!
        # current mth data
        _curr_mth_df = _df.filter(
            (_df["AS_OF_DATE"] == CURR_DATE) &
            (reduce(lambda x, y: x.__and__(y), [_df[c]==row[c] for c in _curve_identifiers]))
        ).with_columns(
            *[pl.col(i).__mul__(pl.lit(100.)) for i in _forward_months_cols], # multiply each MON{-i} col by 100, scale it in %
        ).sort(
            ["DATASET_ID", "AS_OF_DATE"] + _curve_identifiers + ["curve_tenor_days"], descending=False # sort in ascending curve_tenor
        ).select(
            ["DATASET_ID", "AS_OF_DATE"] + _curve_identifiers + ["CURVE_TENOR"] + _forward_months_cols + _metadata_Cols
        ).to_pandas() # in pandas from now onwards
        # prior mth data
        _prior_mth_df = _df.filter(
            (_df["AS_OF_DATE"] == PRIOR_DATE) &
            (reduce(lambda x, y: x.__and__(y), [_df[c]==row[c] for c in _curve_identifiers]))
        ).with_columns(
            *[pl.col(i).__mul__(pl.lit(100.)) for i in _forward_months_cols], # multiply each MON{-i} col by 100, scale it in %
        ).sort(
            ["DATASET_ID", "AS_OF_DATE"] + _curve_identifiers + ["curve_tenor_days"], descending=False # sort in ascending curve_tenor
        ).select(
            ["DATASET_ID", "AS_OF_DATE"] + _curve_identifiers + ["CURVE_TENOR"] + _forward_months_cols + _metadata_Cols
        ).to_pandas() # in pandas from now onwards
        assert _curr_mth_df.shape == _prior_mth_df.shape, f"There are differences in the current & prior DataFrames for {(scen, curve, ccy)}!"
        offset = 1  # Shift by 1 column (e.g., M-24 with M-23)
        if len(_forward_months_cols) > offset:
            # Shift the prior month's columns by the offset
            shifted_prior_cols = _forward_months_cols[offset:]  # Exclude the first `offset` columns
            shifted_curr_cols = _forward_months_cols[:-offset]  # Exclude the last `offset` columns
            # print(len(shifted_prior_cols))
            # print(len(shifted_curr_cols))
            # Take the difference with the offset
            _abs_diff_idx = ["DATASET_ID"] + _curve_identifiers + ["CURVE_TENOR"] + ["SCN_NAME"]
            _abs_diff_df = pd.DataFrame(
                np.abs(_curr_mth_df[shifted_curr_cols].values - _prior_mth_df[shifted_prior_cols].values),
                index=_curr_mth_df[_abs_diff_idx].values,
                columns=shifted_curr_cols,  # Use the current month's column names for the result
            )
           
            _shifted_curr_values = _curr_mth_df[shifted_curr_cols]
            _shifted_prior_values = _prior_mth_df[shifted_prior_cols]
            _filtered_diff_df = _abs_diff_df[_abs_diff_df > 0.7].stack().reset_index()
            _filtered_diff_df.columns = ['Index', 'Column', 'AbsDiff']
            _filtered_diff_df['Shifted_Current_Month'] = _shifted_curr_values.stack().reindex(_filtered_diff_df.set_index(['Index', 'Column']).index).values
            _filtered_diff_df['Shifted_Prior_Month'] = _shifted_prior_values.stack().reindex(_filtered_diff_df.set_index(['Index', 'Column']).index).values
            _filtered_diff_df.to_excel('variance_sheet.xlsx', index=False)
           
            #print(_abs_diff_df)
            _abs_diff_df.index = pd.MultiIndex.from_tuples(
                [tuple(i) for i in _abs_diff_df.index.values],
                names=_abs_diff_idx,
            )
        else:
            raise ValueError(f"Offset {offset} is too large for the number of forward month columns.")
       
        if find_base_vs_shock:
            print(_abs_diff_df.columns)        
        
        combined_column_names = [
            f"{curr_col}Feb-{prior_col}Jan" for curr_col, prior_col in zip(shifted_curr_cols, shifted_prior_cols)
        ]
 
        # Create the DataFrame with the combined column names
        _combined_abs_df = pd.DataFrame(
            np.where(_abs_diff_df.values > abs_tolerance_in_pct, _abs_diff_df.values, None),
            index=_abs_diff_df.index,
            columns=combined_column_names,  # Use the combined column names
        ).stack().dropna().to_frame()
        
        _combined_abs_df.columns = [f"AbsDiff>{abs_tolerance_in_pct}%"]
        
        _combined_abs_df.index = pd.MultiIndex.from_tuples(
            [(*i[:-1], i[-1]) for i in _combined_abs_df.index.values],
            names=_abs_diff_idx + ["FORWARD_MTH"],
        )
        
        print("Absolute Difference DataFrame (_abs_diff_df):")
        print(_abs_diff_df.head())
        non_mon_columns = [f"MON{i}" for i in range(-24, 0)]
        
        print("printing if any values")
        
                 
        _combined_checks_above_tol_list.append(_combined_abs_df) ## append those points DataFrame which exceed abs tolerances
        if (_abs_diff_df.values > 0.7).any():  
            _write_mode = 'a' if output_filename.exists() else 'w'
            kwargs = {} if _write_mode == 'w' else {"if_sheet_exists": "new"}
            ## Export the current mth, prior mth and abs-diff to the output file
            with pd.ExcelWriter(output_filename, engine="openpyxl", mode=_write_mode, **kwargs) as writer:
                # Concatenate _curr_mth_df and _prior_mth_df
                stacked_df = pd.concat(
                    [_curr_mth_df.assign(Source="Current Month"), _prior_mth_df.assign(Source="Prior Month")],
                    axis=0,  # Stack rows
                    ignore_index=True  # Reset the index
                )
   
                # Write the stacked DataFrame to a single sheet
                _ = stacked_df.to_excel(
                    writer,
                    sheet_name=truncate_sheet_name_length(f"{scen}_{curve}_{ccy}_Stacked", curve, max_length=31),
                    index=False  # Do not include the index in the Excel sheet
                )
   
                print(f"Stacked data written to sheet: {truncate_sheet_name_length(f'{scen}_{curve}{ccy}_Stacked', curve, max_length=31)}")
                _ = _abs_diff_df.reset_index().to_excel(
                    writer,
                    sheet_name=truncate_sheet_name_length( f"{scen}_{curve}_{ccy}_ABS-DIFF", curve, max_length=31 ),
                    index=False, # no-index rows
                )
            
            amber_fill = PatternFill(start_color="FFFFBF00", end_color="FFFFBF00", fill_type="solid")
            tabfill=Color('FFFFBF00')
            none_red_fill=PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    # Specify the sheet names
            sheet_names = truncate_sheet_name_length( f"{scen}_{curve}_{ccy}_ABS-DIFF", curve, max_length=31 )
            # Load the workbook
            workbook_path = output_filename
            wb = load_workbook(workbook_path)
            
            if sheet_names in wb.sheetnames:
                sheet = wb[sheet_names]
                # Iterate through all rows and columns in the sheet
                for row in sheet.iter_rows():
                    for cell in row:
                        # Check if the cell value is 0
                        if cell.value == 0:
                            # Apply Amber color fill
                            cell.fill = amber_fill
                            sheet.sheet_properties.tabColor=tabfill
                        elif cell.value == None:
                            cell.fill = none_red_fill
                       
            # Save the updated workbook
            wb.save(workbook_path)
               
        if (_sum_check:=(_abs_diff_df.values > abs_tolerance_in_pct).astype(int).sum()) > 0.7: # TODO: This is really cumbersome
            print(f"{scen}_{curve}_{ccy}_ABS-DIFF have {_sum_check} points outside of abs_tolerance_in_pct={abs_tolerance_in_pct}")
            wb = load_workbook(output_filename)
            ws = wb[truncate_sheet_name_length( f"{scen}_{curve}_{ccy}_ABS-DIFF", curve, max_length=31 )]
            ws.sheet_properties.tabColor = "00FFFF08"
            highlight_fill = PatternFill(start_color="00FFFF08", end_color="00FFFF08", fill_type="solid") # Yellow color
            # for a Dataframe as matrix, incl. of index columns. Note that the header counts as additional + 1 row
            #print(_abs_diff_df.columns)
           
            _abs_diff_df_2 = _abs_diff_df.reset_index()
            _abs_diff_df_2.drop(columns=['SCN_NAME'], axis=1, inplace=True)
            
            _orig_row, _orig_col = _abs_diff_df_2.shape
           
            _start_of_mon = _abs_diff_df_2.columns.values.tolist().index(_abs_diff_df.columns[0])
            _end_of_mon = _start_of_mon + _abs_diff_df.shape[1]
            # print(_start_of_mon, _end_of_mon)
            X, Y = np.meshgrid(np.arange(_orig_col), np.arange(_orig_row + 1), indexing='xy')
           
            # starting from 2nd-row onwards, check for fields which exceeds abs tolerance
            # TODO: This is a real code smell
            X_1=   X[1:, _start_of_mon:_end_of_mon].copy()
            Y_1 =  Y[1:, _start_of_mon:_end_of_mon].copy()
            
           
            assert (X_1.shape == _abs_diff_df.values.shape) and (Y_1.shape == _abs_diff_df.values.shape)
            X_1 = np.where(_abs_diff_df.values > abs_tolerance_in_pct, X_1, np.inf)
            Y_1 = np.where(_abs_diff_df.values > abs_tolerance_in_pct, Y_1, np.inf)
            XY_idx = pd.DataFrame( X_1 ).replace([np.inf, -np.inf], np.nan).stack(future_stack=True).dropna().to_frame()  
            XY_idx.columns = ["i"]  
            XY_idx["j"] = pd.DataFrame( Y_1 ).replace([np.inf, -np.inf], np.nan).stack(future_stack=True).dropna()
            for i, j in XY_idx[["i", "j"]].values:
                ws.cell(row=j+1, column=i+2).fill = highlight_fill # note that in the openpyxl, the 1st-index is 1, not 0
            wb.save(output_filename)
            wb.close()
        else:
            print(f"{scen}_{curve}_{ccy}_ABS-DIFF are within abs_tolerance_in_pct={abs_tolerance_in_pct}")
       
    _write_mode = 'a' if output_filename.exists() else 'w'
    kwargs = {} if _write_mode == 'w' else {"if_sheet_exists": "new"}
    
    _combined_checks_above_tol_list_df = pd.concat(_combined_checks_above_tol_list, axis=0).reset_index()    
    # Check if there are any values greater than 0 in the ABS-DIFF DataFrame
    scenario_id_noval=[]
    scenario_id_withval=[]
    curve_codes_noval=[]
    curve_codes_withval=[]
   
    if (_abs_diff_df.values > 0.7).any():
        
        scenario_id_withval.append(_scenario_ids)
        curve_codes_withval.append(_curve_codes)
        no_values_message = pd.DataFrame([
            {"Category":"Summary:", "Details": f"{_combined_abs_df.count()} records found greater than threshold"},
            {"Category":"Scenarios checked:", "Details":scenario_id_withval},
            {"Category":"Curve Codes:", "Details": curve_codes_withval}])
 
        
       
    else:
        # If there are no values greater than 0, create a VARIANCE_SUMMARY sheet with a message
        scenario_id_noval.append(_scenario_ids)
        curve_codes_noval.append(_curve_codes)
        no_values_message = pd.DataFrame([
            {"Message": ["No values above the threshold were found."]},
            {"Category":"Scenarios checked:", "Details":scenario_id_noval},
            {"Category":"Curve Codes:", "Details": curve_codes_noval}
        ])
        
    scn_ids=_combined_checks_above_tol_list_df["SCN_ID"].astype('int64')
    final_scn_ids=_df.filter(pl.col("SCN_ID").is_in(scn_ids)).select("SCN_NAME").unique().to_series().to_list()
    final_curve_code=_df.filter(pl.col("CURVE_CODE").is_in(_combined_checks_above_tol_list_df["CURVE_CODE"])).select("CURVE_CODE").unique().to_series().to_list()
    final_ccy_code=_df.filter(pl.col("CCY_CODE").is_in(_combined_checks_above_tol_list_df["CCY_CODE"])).select("CCY_CODE").unique().to_series().to_list()
    
    summary_data={"Threshold":[f"Abs({abs_tolerance_in_pct}%)"], 
                  "Records found greater than threshold":[len(_combined_checks_above_tol_list_df)],
                  "Scenarios":[", ".join(map(str, final_scn_ids))],
                "Curve_codes_with_abs_diff_greater_than_0.7": [", ".join(map(str, final_curve_code))],
                 "CCY_CODE":[", ".join(map(str, final_ccy_code))]}
    # write results to Excel
    
    _combined_checks_above_tol_list_df = pd.concat(_combined_checks_above_tol_list, axis=0).reset_index()
    with pd.ExcelWriter(output_filename, engine="openpyxl", mode="a", if_sheet_exists='overlay') as writer:
        if not _combined_checks_above_tol_list_df.empty:
            
            _combined_checks_above_tol_list_df.to_excel(
                    writer,
                    sheet_name="VARIANCE_SUMMARY",  
                    index=False  # Do not include the index in the Excel sheet
                )
        else:
            pd.DataFrame({"DATASET_ID":["No Variance found"]}).to_excel(writer,sheet_name="VARIANCE_SUMMARY", index=False)
        # write summary to the right side
        pd.DataFrame(summary_data).T.to_excel(writer, sheet_name= "VARIANCE_SUMMARY", startcol = 11, startrow=1, header=False)
                                             
    wb = load_workbook(output_filename)
    wb.move_sheet("VARIANCE_SUMMARY", -(len(wb.sheetnames)-1))
    ws = wb["VARIANCE_SUMMARY"]
    ws.sheet_properties.tabColor = "00000000"    
    print("VARIANCE_SUMMARY sheet created.")
   
    
    for ws in wb.worksheets:
        ws.sheet_view.zoomScale = 75
    highlight_fill = PatternFill(start_color="00FFFF08", end_color="00FFFF08", fill_type="solid")
    if not _combined_checks_above_tol_list_df.count().empty:
        sheet_names = truncate_sheet_name_length( f"{scen}_{curve}_{ccy}_ABS-DIFF", curve, max_length=31 )
        if sheet_names in wb.sheetnames:
            sheet=wb[sheet_names]
            for row in sheet.iter_rows():
                for cell in row:
                    try:
                        value=float(cell.value) if cell.value is not None else None
                        if value is not None and value>0.7:
                            cell.fill=highlight_fill
                            # print(f"Changed color for cell {cell.coordinate} with value {value}")
                    except(ValueError, TypeError):
                        continue
    wb.save(output_filename)
    wb.close()    
    return _abs_diff_df

def merge_parquet_files(folder_path):
    """Merge all Parquet files in a given folder into a single DataFrame."""
    parquet_files = glob.glob(f"{folder_path}/*.csv")
    if not parquet_files:
        raise FileNotFoundError(f"No Parquet files found in the folder: {folder_path}")
    
    dataframes = []
    for file in parquet_files:
        try:
            # df = pd.read_parquet(file)
            df = pd.read_csv(file, parse_dates=["AS_OF_DATE"])

            
            if not df.empty:
                dataframes.append(df)
        except Exception as e:
            print(f"Error reading {file}: {e}")
    
    if not dataframes:
        raise ValueError(f"All Parquet files in {folder_path} are empty or invalid.")
    
    return pd.concat(dataframes, ignore_index=True)

def main():
    import argparse
    import time
    #print('7:4am')
    parser = argparse.ArgumentParser(
        prog='ReaderIRCurve',
        description='This exports the IRCurve differences between Current Month vs Prior Month.',
        epilog='Check Ruby Close',
    )
    ## TODO: to add parser arguments to the method reader_and_export()
    parser.add_argument("--current_month", required=True, help="Path to the folder containing Parquet files for the current month.")
    parser.add_argument("--prior_month", required=True, help="Path to the folder containing Parquet files for the prior month.")    
#     parser.add_argument('--input_data_location', type=str) ## full path unix-style (ie. as /) for the parquet IRCurve file
    parser.add_argument("--output_file", required=False, help="Path to save the merged Parquet file (optional).")


    parser.add_argument('--scenario_data_location', type=str) ## full path unix-style (ie. as /) for the Excel Scenario file
    parser.add_argument('--scenario_id', nargs='+', type=int) # either one or more scenario_id as integer
    parser.add_argument('--curve_code', nargs='+', type=str) # either one or more curve_code as str
    parser.add_argument('--ccy_code', nargs='+', type=str) # either one or more ccy_code as str
    parser.add_argument('--output_filename', type=str) ## full path unix-style (ie. as /) for the final output filename
    parser.add_argument('--abs_tolerance_in_pct', nargs='?', default=0.7, type=float) ## the absolute tolerance difference in %
    ## TODO: add the forward_months_cols as argparse argument, when time permits
    args = parser.parse_args()
    print("Start processing the comparisons.....")
    start = time.perf_counter()
    timestamp=datetime.now().strftime("%Y%m%d_%H%M%S")
    # Merge Parquet files for current and prior months
    print(f"Merging Parquet files from current month folder: {args.current_month}")
    current_month_df = merge_parquet_files(args.current_month)
    
    print(f"Merging Parquet files from prior month folder: {args.prior_month}")
    prior_month_df = merge_parquet_files(args.prior_month)
    
    # Combine the two DataFrames
    merged_df = pd.concat([current_month_df, prior_month_df], ignore_index=True)
    print("Successfully merged current and prior month data.")
    
    # Save to output file if specified
    if args.output_file:
        merged_df.to_parquet(args.output_file, index=False)
        print(f"Merged DataFrame saved to {args.output_file}")
    else:
        print("No output file specified. DataFrame is ready for further processing.")
            
    output_filename= Path(f"{args.output_filename}/test_variance_report_{args.curve_code}_{timestamp}.xlsx")
    ruby_data = RubyProductionData(
        input_data_location = Path(args.output_file),
        scenario_data_location = Path(args.scenario_data_location),
    )
    # _ = reader_and_export(
    #     ruby_data= ruby_data,
    #     scenario_id = args.scenario_id,
    #     curve_code = args.curve_code,
    #     ccy_code = args.ccy_code,
    #     output_filename = output_filename,
    #     abs_tolerance_in_pct = float(args.abs_tolerance_in_pct),
    # )
    _ = optimized_reader_export(
        _df = ruby_data.input_data_polars,
        scenario_ids = args.scenario_id,
        curve_codes = args.curve_code,
        ccy_codes = args.ccy_code,
        output_filename = output_filename,
        abs_tolerance_in_pct = float(args.abs_tolerance_in_pct),
    )
    print(f"Finished processing the comparisons, Time-taken={time.perf_counter() - start:.2f} sec.", )
if __name__ == "__main__":
    main()