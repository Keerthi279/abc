import polars as pl
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, NamedStyle, Color
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, Tuple, Union
from functools import cached_property
import re
from itertools import product
from datetime import datetime
from functools import reduce


class MetaSingleton(type):
    """Classic MetaSingleton creates an instance only if there is no instance created so far;
    otherwise, it will return the instance that is already created."""
    
    _instances = {}
    def __call__(cls, *args, **kwargs):
        if cls not in cls._instances:
            cls._instances[cls] = super().__call__(*args, **kwargs)
        return cls._instances[cls]

@dataclass(frozen=True)
class RubyProductionData(metaclass=MetaSingleton):
    input_data_location: Path
    scenario_data_location: Path
    columns_for_sorting: Tuple[str] = field(
        default_factory= lambda: ["SCN_ID","CURVE_DETL_ID","CURVE_MSTR_ID","CURVE_TENOR","AS_OF_DATE"]
    )

    @cached_property
    def scen_data_polars(self):
        df = pl.read_excel(self.scenario_data_location, engine="openpyxl").select(
            [pl.col("SCN_ID").alias("SCN_ID_right"), pl.col("SCN_NAME")]  # TODO: validate these headers
        )

        return df

    @cached_property
    def input_data_polars(self):
        # TODO: validate that these column headers are present
        df = (
            pl.read_parquet(
                self.input_data_location
                ).sort(
                    self.columns_for_sorting, descending=False
                    ).with_columns(
                pl.col("CURVE_TENOR").replace_strict(
                    {"ZERO": 0, "O/N": 1, "7D": 7, "14D": 14, **{f"{i}M": i * 30 for i in range(1, 600)}}
                ).cast(pl.Int32).alias("curve_tenor_days")
            ).join(
                self.scen_data_polars,
                left_on="SCN_ID", right_on="SCN_ID_right"
            )
        )

        # TODO: validate these column headers
        new_cols_order = ["SCN_ID", "SCN_NAME", "AS_OF_DATE", "CURVE_CODE", "CCY_CODE","curve_tenor_days", "CURVE_TENOR"] + \
                         [f"MON{i}" for i in range(-24, 61)] + \
                         ["CURVE_DETL_ID", "CURVE_MSTR_ID", "DATASET_ID", "CURVE_DESC", "UNITS", "ORIGL_FLAG", "CURVE_FLOOR", "CURVE_CNTRY", "FWD_SPRD_BOOL",
                          "CREATED_BY", "CREATE_TS", "SOURCE", "SOURCE_DETAILS"]
        df = df.select(new_cols_order)

        return df
    
def truncate_sheet_name_length(original_str: str, field_to_truncate: str, max_length: int = 31) -> str:
    if len(original_str) <= max_length:
        return original_str  # If less than max length, return original string
    _remainder_str = original_str.replace(field_to_truncate, '')
    _residual_len = max_length - len(_remainder_str)
    
    return re.sub(f"({re.escape(field_to_truncate)})", field_to_truncate[:_residual_len], original_str)

def reader_and_export(
    ruby_data: RubyProductionData,
    scenario_id: Union[int, Tuple[int]],
    curve_code: Union[str, Tuple[str]],
    ccy_code: Union[str, Tuple[str]],
    output_filename: Path,
    abs_tolerance_in_pct: float = 0,
    forward_months_cols: Tuple[str] = tuple([f"M{i}" for i in range(-24, 61)]),
    find_base_vs_shock: bool = False
) -> None:

    # TODO: From Ruby data, the accepted headers are MON-24, MON-23, ...

    _forward_months_cols = ["MON{used_int}".format(used_int=re.findall(r'[\-|\+]*\d+', i)[0]) for i in forward_months_cols]
    _scenario_ids = [scenario_id] if isinstance(scenario_id, int) else list(scenario_id)
    _curve_codes = [curve_code] if isinstance(curve_code, str) else list(curve_code)
    _ccy_codes = [ccy_code] if isinstance(ccy_code, str) else list(ccy_code)

    _df = ruby_data.input_data_polars


    _dates_used = pd.to_datetime(_df[["AS_OF_DATE"]].unique().to_pandas().values.flatten())
    assert len(_dates_used) == 2, "Make sure there are only current and prior months!"
    CURR_DATE, PRIOR_DATE = max(_dates_used), min(_dates_used)
    _curve_identifiers = ["SCN_ID", "CURVE_CODE", "CCY_CODE"]
    _metadata_Cols = ["CURVE_MSTR_ID", "SCN_NAME", "ORIGL_FLAG", "CURVE_FLOOR", "CREATED_BY", "CREATE_TS", "SOURCE", "SOURCE_DETAILS"]
    _combined_checks_above_tol_list = []
    if output_filename.exists():  # TODO: Need to check if it is being opened
        os.remove(output_filename)

    wb = Workbook()

    # Rename the default sheet
    ws = wb.active
    ws.title = "VARIENCE_SUMMARY"


    # Save the workbook
    wb.save(output_filename)

    print("Excel file with multiple empty sheets created successfully!")
    for scen, curve, ccy in product(_scenario_ids, _curve_codes, _ccy_codes):
        row = {"SCN_ID": scen, "CURVE_CODE": curve, "CCY_CODE": ccy}  # TODO: refactor this unnecessary assignment

        # current month data
        _curr_mth_df = _df.filter(
            (_df["AS_OF_DATE"] == CURR_DATE) &
            (reduce(lambda x, y: x & y, [_df[c] == row[c] for c in _curve_identifiers]))
        ).with_columns(
            [pl.col(i).mul(pl.lit(100.0)) for i in _forward_months_cols]  # multiply MON{-i} by 100 (percent scaling)
        ).sort(
            ["DATASET_ID", "AS_OF_DATE"] + _curve_identifiers + ["curve_tenor_days"], descending=False
        ).select(
            ["DATASET_ID", "AS_OF_DATE"] + _curve_identifiers + ["CURVE_TENOR"] + _forward_months_cols + _metadata_Cols
        ).to_pandas()  # in pandas from now onwards
        # prior month data
        _prior_mth_df = _df.filter(
            (_df["AS_OF_DATE"] == PRIOR_DATE) &
            (reduce(lambda x, y: x & y, [_df[c] == row[c] for c in _curve_identifiers]))
        ).with_columns(
            [pl.col(i).__mul__(pl.lit(100.0)) for i in _forward_months_cols]  # multiply each MON{-i} col by 100
        ).sort(
            ["DATASET_ID", "AS_OF_DATE"] + _curve_identifiers + ["curve_tenor_days"], descending=False
        ).select(
            ["DATASET_ID", "AS_OF_DATE"] + _curve_identifiers + ["CURVE_TENOR"] + _forward_months_cols + _metadata_Cols
        ).to_pandas()  # in pandas from now onwards
        assert _curr_mth_df.shape == _prior_mth_df.shape, f"There are differences in the current & prior DataFrames for {(scen, curve, ccy)}!"

        # Define the offset for the calculation
        
        
        
        
        
        
        offset = 1  # Shift by 1 column (e.g., M-24 with M-23)

        print(_curr_mth_df.shape)
        print(_prior_mth_df.shape)

        # Ensure the number of columns in _forward_months_cols allows for the offset
        if len(_forward_months_cols) > offset:
            # Shift the prior month's columns by the offset
            shifted_prior_cols = _forward_months_cols[offset:]  # Exclude first `offset` columns
            shifted_curr_cols = _forward_months_cols[:-offset]  # Exclude last `offset` columns
            print(len(shifted_prior_cols))
            print(len(shifted_curr_cols))
            # Take the difference with the offset
            _abs_diff_idx = ["DATASET_ID"] + _curve_identifiers + ["CURVE_TENOR"] + ["SCN_NAME"]
            _abs_diff_df = pd.DataFrame(
                np.abs(_curr_mth_df[shifted_curr_cols].values - _prior_mth_df[shifted_prior_cols].values),
                index=_curr_mth_df[_abs_diff_idx].values,
                columns=shifted_curr_cols
            )




            _shifted_curr_values = _curr_mth_df[shifted_curr_cols]
            _shifted_prior_values = _prior_mth_df[shifted_prior_cols]
            _filtered_diff_df = _abs_diff_df[_abs_diff_df > 0.7].stack().reset_index()
            _filtered_diff_df.columns = ['Index', 'Column', 'AbsDiff']
            _filtered_diff_df['Shifted Current Month'] = _shifted_curr_values.stack().reindex(_filtered_diff_df.set_index(['Index', 'Column']).index).values
            _filtered_diff_df['Shifted Prior Month'] = _shifted_prior_values.stack().reindex(_filtered_diff_df.set_index(['Index', 'Column']).index).values
            _filtered_diff_df.to_excel("variance_sheet.xlsx", index=False)


            _abs_diff_df.index = pd.MultiIndex.from_tuples(
                [tuple(i) for i in _abs_diff_df.index.values],
                names=_abs_diff_idx,
            )
        else:
            raise ValueError(f"Offset {offset} is too large for the number of forward month columns.")




    if find_base_vs_shock:
        print(_abs_diff_df.columns)







        # Create a mapping of column names to represent the combination of current and prior months
        combined_column_names = [
            f"{curr_col}↔{prior_col}" for curr_col, prior_col in zip(shifted_curr_cols, shifted_prior_cols)
        ]

        # Create the DataFrame with the combined column names
        _combined_abs_df = pd.DataFrame(
            np.where(_abs_diff_df.values > abs_tolerance_in_pct, _abs_diff_df.values, None),
            index=_abs_diff_df.index,
            columns=combined_column_names,
        ).stack().dropna().to_frame()
        print(_combined_abs_df.columns, "here")

        # Rename the column to indicate the absolute difference exceeding the threshold
        _combined_abs_df.columns = [f"AbsDiff>{abs_tolerance_in_pct}%"]
        print(_combined_abs_df.columns, "here2")

        print(_combined_abs_df.index)


        # Update the index to include the FORWARD_MTH information
        _combined_abs_df.index = pd.MultiIndex.from_tuples(
            [(i[:-1], i[-1]) for i in _combined_abs_df.index.values],
            names=_abs_diff_idx + ["FORWARD_MTH"]
        )

        # Assign the curr mth values which exceed abs tol
        # print(_curr_mth_df[shifted_curr_cols].values)
        # _combined_abs_df["2025-02"] = pd.DataFrame(
        #     np.where(_abs_diff_df.values > abs_tolerance_in_pct, _curr_mth_df[shifted_curr_cols].values, None),
        #     index=_abs_diff_df.index,
        #     columns=shifted_curr_cols
        # ).stack().reindex(_combined_abs_df.index).dropna()




        print("Absolute Difference DataFrame (_abs_diff_df):")
        print(_abs_diff_df.head())

        non_mon_columns = [f"MON{i}" for i in range(0, 60)]
        print(non_mon_columns)

        _abs_diff_df = _abs_diff_df.drop(columns=non_mon_columns)
        print(_abs_diff_df.columns)

        print("Current Month DataFrame (_curr_mth_df):")
        print(_curr_mth_df.head())

        print("Prior Month DataFrame (_prior_mth_df):")
        print(_prior_mth_df.head())

        print("Combined DataFrame (_combined_abs_df):")
        print(_combined_abs_df.head())

        print(_prior_mth_df[_abs_diff_df.columns].columns)
        print(_prior_mth_df[shifted_prior_cols].values)
        print("printing if any values")



















        _combined_checks_above_tol_list.append(_combined_abs_df)  # append DataFrame exceeding tolerances
        if (_abs_diff_df.values > 0).any():
            write_mode = "a" if output_filename.exists() else "w"
            kwargs = {} if write_mode == "w" else {"if_sheet_exists": "new"}

            with pd.ExcelWriter(output_filename, engine="openpyxl", mode=write_mode, **kwargs) as writer:
                
                stacked_df = pd.concat(
                    [_curr_mth_df.assign(Source="Current Month"), _prior_mth_df.assign(Source="Prior Month")],
                    axis=0,  # stack rows
                    ignore_index=True
                )

                # Write stacked DataFrame to single sheet
                stacked_df.to_excel(
                    writer,
                    sheet_name=truncate_sheet_name_length(f"{scen}_{curve}_{ccy}_Stacked", curve, max_length=31),
                    index=False
                )

                print(f"Stacked data written to sheet: {truncate_sheet_name_length(f'{scen}_{curve}_{ccy}_Stacked', curve, max_length=31)}")
                _ = _abs_diff_df.reset_index().to_excel(
                    writer,
                    sheet_name=truncate_sheet_name_length(f"{scen}_{curve}_{ccy}_ABS-DIFF", curve, max_length=31),
                    index=False
                )


        # Apply formatting with openpyxl if needed
        amber_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
        tabfill = Color('FFFFBF00')
        none_red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        sheet_names = truncate_sheet_name_length(f"{scen}_{curve}_{ccy}_ABS-DIFF", curve, max_length=31)

        workbook_path = output_filename
        wb = load_workbook(workbook_path)





        if sheet_names in wb.sheetnames:
            sheet = wb[sheet_names]

            for row in sheet.iter_rows():
                for cell in row:

                    if cell.value == 0:

                        cell.fill = amber_fill
                        sheet.sheet_properties.tabColor = tabfill
                    elif cell.value == None:
                        cell.fill = none_red_fill


        wb.save(workbook_path)

    if (_sum_check:=(_abs_diff_df.values > abs_tolerance_in_pct).astype(int).sum()) > 0: # TODO: This is really cumbersome
        print(f"{scen}_{curve}_{ccy} ABS-DIFF have {_sum_check} points outside of abs_tolerance_in_pct={abs_tolerance_in_pct}")
        wb = load_workbook(output_filename)
        ws = wb[truncate_sheet_name_length( f"{scen}_{curve}_{ccy}_ABS-DIFF", curve, max_length=31 )]
        ws.sheet_properties.tabColor = "00FFFF00"
        highlight_fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid") # Yellow color
        # For a dataframe as matrix, incl. of index columns. Note that the header counts as additional + 1 row
        #print(_abs_diff_df.columns)

        _abs_diff_df = _abs_diff_df.reset_index()
        _abs_diff_df_2 = _abs_diff_df.drop(columns=["SCN_NAME"], axis=1, inplace=True)
        print(_abs_diff_df_2.shape)


        _orig_row, _orig_col = _abs_diff_df_2.shape

        _start_of_mon = _abs_diff_df_2.colums.values.tolist().index(_abs_diff_df.colums[0])
        _end_of_mon = _start_of_mon + _abs_diff_df.shape[1]
        print(_start_of_mon, _end_of_mon)
        X, Y = np.meshgrid(np.arange(_orig_col), np.arange(_orig_row + 1), indexing="xy")

        # starting from 2nd-row onwards, check for fields which exceeds abs tolerance
        # T000: This is a real code smell
        X_1 = X[1:, _start_of_mon: _end_of_mon].copy()
        Y_1 = Y[1:, _start_of_mon: _end_of_mon].copy()






        

        assert (X_1.shape == _abs_diff_df.values.shape) and (Y_1.shape == _abs_diff_df.values.shape)
        X_1 = np.where(_abs_diff_df.values > abs_tolerance_in_pct, X_1, np. inf)
        Y_1 = np.where(_abs_diff_df.values > abs_tolerance_in_pct, Y_1, np. inf)
        XY_idx = pd.vatarrane( X_1 ).replace([np.inf, -np.inf], np.nan) .stack(future_stack=True).dropna().to_frame()
        XY_idx.columns = ["i"]
        XY_idx["j"] = pd.DataFrame( Y_1 ).replace([np.inf, -np.inf], np.nan).stack(future_stack=True) .dropna()
        for i, j in XY_idx[["i", "j"]].values:
            ws.cell(row=j+1, column=i+2).fill = highlight_fill # note that in the openpyxl, the 1st-index is 1, not @
        wb. save(output_filename)
        wb.close()
    else:
        print(f"{scen}_{curve}_{ccy}_A8S-DIFF are within abs tolerance_in_pct={abs_tolerance_in_pct}")

    _write_mode = 'a' if output_filename.exists() else "w"
    kwargs = {} if _write_mode == "w" else {"if sheet exists": "new"}

    print(len(_combined_checks_above_tol_list))
    for i in range(len(_combined_checks_above_tol_list)):
        print(_combined_checks_above_tol_list[i].columns)

    _combined_checks_above_tol_list_df = pd.concat(_combined_checks_above_tol_list, axis=0).reset_index()

    scenario_id_noval = []
    scenario_id_withval = []
    curve_codes_noval = []
    curve_codes_withval = []

    if (_abs_diff_df.values > 0).any():







        scenario_id_withval.append(_scenario_ids)
        curve_codes_withval.append(_curve_codes)

        no_values_message = pd.DataFrame([
            {"Category": "Summary", "Details": f"{_combined_abs_df.count()} records found greater than threshold"},
            {"Category": "Scenarios checked", "Details": scenario_id_withval},
            {"Category": "Curve Codes", "Details": curve_codes_withval}
        ])











    else:
        # If there are no values greater than 0, create a VARIENCE_SUMMARY sheet with a message
        scenario_id_noval.append(_scenario_ids)
        curve_codes_noval.append(_curve_codes)
        no_values_message = pd.DataFrame([
            {"Message": ["No values above the threshold were found."]},
            {"Category": "Scenarios checked:", "Details": scenario_id_noval},
            {"Category": "Curve Codes", "Details": curve_codes_noval}
        ])










    filtered_df = _abs_diff_df[_abs_diff_df.gt(abs_tolerance_in_pct).any(axis=1)]

    # Get SCN_IDs and associated metadata
    filtered_ids = filtered_df.index.get_level_values("SCN_ID").unique().tolist()

    # Prepare the summary output
    summary_data = {
        "Threshold": [f"Abs ({abs_tolerance_in_pct}%)"],
        "Records found greater than threshold": [len(filtered_df)],
        "Scenarios": [sorted(set(_df.filter(pl.col("SCN_ID").is_in(filtered_ids))["SCN_NAME"].unique().to_list()))],
        "Curve": [sorted(set(_df.filter(pl.col("SCN_ID").is_in(filtered_ids))["CURVE_CODE"].unique().to_list()))],
        "CCY": [sorted(set(_df.filter(pl.col("SCN_ID").is_in(filtered_ids))["CCY_CODE"].unique().to_list()))]
    }

    # Write results to Excel
    with pd.ExcelWriter(output_filename, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        if not filtered_df.empty:
            filtered_df.to_excel(writer, sheet_name="VARIENCE_SUMMARY", index=False)
        else:
            pd.DataFrame({"DATASET_ID": ["No variance found"]}).to_excel(writer, sheet_name="VARIENCE_SUMMARY", index=False)

        # Write summary at right side
        pd.DataFrame(summary_data).T.to_excel(writer, sheet_name="VARIENCE_SUMMARY", startcol=11, startrow=1, header=False)
    
    with pd.ExcelWriter(output_filename, engine="openpyxl", mode="a", if_sheet_exists='replace') as writer:
        no_values_message.to_excel(
            writer,
            sheet_name="VARIENCE_SUMMARY",
            index=False  # Do not include the index in the Excel sheet
        )

    wb = load_workbook(output_filename)
    wb.move_sheet("VARIENCE_SUMMARY", -(len(wb.sheetnames)-1))
    ws = wb["VARIENCE_SUMMARY"]
    ws.sheet_properties.tabColor = "00000000"
    print("VARIENCE_SUMMARY sheet created.")


        













    for ws in wb.worksheets:
        ws.sheet_view.zoomScale = 75
    highlight_fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")
    for row in ws.iter_rows():
        for cell in row:
            try:
                value = float(cell.value)
                if cell.value is not None and value > 0.7:
                    cell.fill = highlight_fill
                    print(f"Changed color for cell {cell.coordinate} with value {value}")
            except (ValueError, TypeError):
                continue

    wb.save(output_filename)
    return _abs_diff_df

def main():
    parser = argparse.ArgumentParser(
        prog="ReaderIRCurve",
        description="This exports the IRCurve differences between Current Month vs Prior Month.",
        epilog="Check Ruby Close",
    )

    # TODO: add parser arguments to the method reader_and_export()
    parser.add_argument("--input_data_location", type=str)  # full path unix-style (ie. as /) for the parquet IRCurve file
    parser.add_argument("--scenario_data_location", type=str)  # full path unix-style (ie. as /) for the Excel Scenario file
    parser.add_argument("--scenario_id", nargs="+", type=int)  # either one or more scenario_id as integer
    parser.add_argument("--curve_code", nargs="+", type=str)  # either one or more curve code as str
    parser.add_argument("--ccy_code", nargs="+", type=str)  # one or more ccy_code as str
    parser.add_argument("--output_filename", type=str)  # full path unix-style (ie. as /) for the final output filename
    parser.add_argument("--abs_tolerance_in_pct", nargs="?", default=0.7, type=float)  # the absolute tolerance difference in %

    # TODO: add the forward_months_cols as argparse argument, when time permits
    args = parser.parse_args()
    print("start processing the comparisons.....")
    start = time.perf_counter()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = Path(f"{args.output_filename}/test_varience_report_{timestamp}.xlsx")

    ruby_data = RubyProductionData(
        input_data_location=Path(args.input_data_location),
        scenario_data_location=Path(args.scenario_data_location),
    )

    _ = reader_and_export(
        ruby_data=ruby_data,
        scenario_id=args.scenario_id,
        curve_code=args.curve_code,
        ccy_code=args.ccy_code,
        output_filename=output_filename,
        abs_tolerance_in_pct=float(args.abs_tolerance_in_pct),
    )

    print(f"Finished processing the comparisons, Time-taken={time.perf_counter() - start:.2f} sec.")

if __name__ == "__main__":
    main()





